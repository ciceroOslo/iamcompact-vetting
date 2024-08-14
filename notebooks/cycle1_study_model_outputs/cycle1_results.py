"""Module that loads the result files from the 1st modelling cycle on import."""

# %%
# Imports
from pathlib import Path
import typing as tp
import pickle
import hashlib

import pyam
import pandas as pd
try:
    import openpyxl
except ModuleNotFoundError:
    raise ModuleNotFoundError(
        'Package `openpyxl` not found. This package is required to load the '
        'result Excel files.'
    )



# %%
# Define a utility function that takes a variable of type T|None for some type
# T and returns either the value of the variable if it is of type T or raises
# a ValueError if the variable is None (i.e., the function is guaranteed to have
# return type T)
TV = tp.TypeVar('TV')
def notnone(x: TV|None) -> TV:
    if x is None:
        raise ValueError('Value is None')
    return x
###END def notnone

# %%
# Create a helper function that takes a function as an argument, and returns
# a modified function which will return the return value of the original
# function if it returns, but catch any exceptions and return them. The function
# should be properly type-annotated, with ParamSpecs to ensure that the
# modified function has the same signature and return type as the original
# function, except that it can return an exception as well as the original
# return type.
PS = tp.ParamSpec('PS')  # Parameter specifiction
RT = tp.TypeVar('RT')  # Return type
def return_exceptions(
    func: tp.Callable[PS, RT]
) -> tp.Callable[PS, RT|Exception]:
    def wrapper(*args: PS.args, **kwargs: PS.kwargs) -> RT|Exception:
        try:
            return func(*args, **kwargs)
        except Exception as e:
            return e
    return wrapper

# %%
# Define a function to open IAMC excel files with multiple sheets
def open_multisheet_iamc(path: Path|str) -> \
        pyam.IamDataFrame|Exception|dict[str, pyam.IamDataFrame|Exception]:
    sheetnames: list[str] = openpyxl.open(path).sheetnames
    if len(sheetnames) == 1:
        return return_exceptions(pyam.IamDataFrame)(path)
    else:
        return {
            sheetname: return_exceptions(pyam.IamDataFrame) \
                (path, sheet_name=sheetname)
            for sheetname in sheetnames
        }

# %%
# Get data files, and create a nested dict of IamDataFrames
# If a cached pickle file exists, load it. Otherwise, load the data files and
# save the dict as a pickle file.
try:
    data_root: Path = Path(__file__).parent
except NameError:
    data_root: Path = Path.cwd()
cache_file: Path = data_root / 'data_dict.pkl'
FORCE_RELOAD: bool = False
pickle_hash: str = 'a5aaeeda88c914a9b55b9763493c50d8'
write_cache: bool = True

if cache_file.exists() and not FORCE_RELOAD:
    # First check that the md5sum hash of the pickle file matches `pickle_hash`.
    # Raise an error if not.
    print(f'Reading cache file {cache_file}, expecting MD5 hash {pickle_hash}')
    with open(cache_file, 'rb') as f:
        cache_file_content: bytes = f.read()
        cache_file_hash = hashlib.md5(cache_file_content).hexdigest()
        if cache_file_hash != pickle_hash:
            raise ValueError(
                f'Hash of pickle file {cache_file} does not match expected hash'
            )
    # Load the pickle file using the content that was already read
    data_dict: dict[str, pyam.IamDataFrame | Exception |
                    dict[str, pyam.IamDataFrame | Exception]] = pickle.loads(
        cache_file_content
    )
    print('Successfully loaded cache file.')
else:
    if not cache_file.exists():
        print(f'Cache file {cache_file} does not exist. Loading data files.')
    elif FORCE_RELOAD:
        print(f'FORCE_RELOAD is set to True. Loading data files.')
    else:
        raise RuntimeError(
            'Cache file exists, but FORCE_RELOAD is not set to True. It should '
            'not be possible to reach this point in the code in this case, '
            'please check for errors in the code.'
        )
    data_dict: dict[str, pyam.IamDataFrame | Exception |
                        dict[str, pyam.IamDataFrame | Exception]] = dict()
    _p: Path
    _idf: pyam.IamDataFrame | Exception | dict[str, pyam.IamDataFrame | Exception]
    relpaths: dict[Path, Path] = {
        _p: _p.relative_to(data_root)
        for _p in data_root.glob('**/*.xlsx')
    }
    _relpath: Path
    _abspath: Path
    for _abspath, _relpath in relpaths.items():
        _idf = open_multisheet_iamc(_abspath)
        data_dict[str(_relpath)] = _idf
    if write_cache:
        if cache_file.exists():
            raise FileExistsError(
                f'Cache file {cache_file} already exists. Please delete it '
                'before writing a new cache file, or set a different name/path '
                'in the code.'
            )
        with open(cache_file, 'wb') as f:
            pickle.dump(data_dict, f)
        # Calculate the hash of the pickle file and print it.
        with open(cache_file, 'rb') as f:
            cache_file_content = f.read()
            cache_file_hash = hashlib.md5(cache_file_content).hexdigest()
            print(
                f'Cache file written to {str(cache_file)} with MD5 hash '
                f'{cache_file_hash}'
            )

# %%
# Flatten the dict by inserting an @ sign in front of tab names for files that
# have multiple tabs
flat_data_dict: dict[str, pyam.IamDataFrame | Exception] = dict()
_pathkey: str
for _pathkey, _idf in data_dict.items():
    if isinstance(_idf, dict):
        for _tabname, _idf_tab in _idf.items():
            flat_data_dict[f'{_pathkey}@{_tabname}'] = _idf_tab
    else:
        flat_data_dict[_pathkey] = _idf

# %%
# Get only the items that are IamDataFrames
flat_data_dict_iamdfs: dict[str, pyam.IamDataFrame] = {
    _relpath: _idf for _relpath, _idf in flat_data_dict.items()
    if isinstance(_idf, pyam.IamDataFrame)
}

# %%
# Try to make a joint IamDataFrame by concatenating all the IamDataFrames from
# `flat_data_dict_iamdfs` except the ones for which they key contains 
# `'nolinksv1'`, but first rename the scenarios in each IamDataFrame by
# appending the first three characters of the key (of the form 'S0n' for some
# digit n) to the scenario name.
all_iamdfs_list: list[pyam.IamDataFrame] = []
_key: str
_iamdf: pyam.IamDataFrame
_scenario: str
for _key, _iamdf in flat_data_dict_iamdfs.items():
    if 'nolinksv1' in _key:
        continue
    _key_first3 = _key[:3]
    _iamdf = notnone(
        notnone(_iamdf).rename(
            scenario={_scenario: f'{_scenario}_{_key_first3}'
                    for _scenario in _iamdf.scenario}
        )
    )
    all_iamdfs_list.append(_iamdf)

# %%
# Now join all the DataFrames. Do it one by one in order to track down where any
# errors occur. CHANGE: Commenting out the code that does it one by one, and
# and instead concatenating all at once, which hopefully will be more
# performant.
def _remove_unnamed0_col(iamdf: pyam.IamDataFrame) -> pyam.IamDataFrame:
    if len(iamdf.extra_cols) > 0:
        if len(iamdf.extra_cols) > 1 or iamdf.extra_cols[0] != 'unnamed: 0':
            raise ValueError(
                'An IamDataFrame contains an unexpected extra index column.'
            )
            # Remove the extra column
        iamdf_orig: pyam.IamDataFrame = iamdf
        iamdf = pyam.IamDataFrame(iamdf._data.droplevel('unnamed: 0'))
    return iamdf

# joint_iamdf: pyam.IamDataFrame = _remove_unnamed0_col(all_iamdfs_list[0])
# for _iamdf in all_iamdfs_list[1:]:
#     _iamdf_orig: pyam.IamDataFrame = _iamdf
#     _iamdf = _remove_unnamed0_col(_iamdf)
#     joint_iamdf = notnone(joint_iamdf.append(_iamdf))

joint_iamdf: pyam.IamDataFrame = pyam.concat(
    [
        _remove_unnamed0_col(_iamdf) for _iamdf in all_iamdfs_list
    ]
)
